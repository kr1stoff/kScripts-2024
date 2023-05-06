#!/usr/bin/env python

import os
import yaml
import logging
import subprocess
import time
import zipfile
import shutil
from subprocess import run
from openpyxl import load_workbook


# 给库添加日志
logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler)

#通用类型#######################################################################
def dict2yaml(src_dict:dict, dst_yaml:str):
    """
    [220601 mxf] dict转成yaml
        src_dict    "任务-样本"大字典
        dst_yaml    输出yaml路径
    """
    os.makedirs(os.path.dirname(dst_yaml), exist_ok=True)
    with open(dst_yaml, "wt", encoding="utf-8", newline="") as gh:
        gh.write(yaml.safe_dump(src_dict, allow_unicode=True))

def dict2sampletable(src_dict, dst_table):
    """
    [220602 mxf] yaml文件转为样本表,新冠NGSWGA流程用
    测试task_id: 8936(PE), 8942(SE)
    """
    gh = open(dst_table, "wt", encoding="utf-8", newline="")
    for sn in src_dict["samples"]: # sample_number
        fq1 = src_dict["samples"][sn]["sequencing_data1"]
        fq2 = src_dict["samples"][sn]["sequencing_data2"]
        outline = "\t".join([sn, fq1, fq2]) + "\n"
        gh.write(outline)
    gh.close()

def excel2bed(excel:str, bed):
    """
    [220610] 新冠流程输入为excel引物位置信息, 需要转成bed跑流程
    参数：
        excel   VENUS excel格式引物位置信息文件
        bed     bed引物位置信息文件
    返回：
        失败 --> ""
        成功 --> bed
    """
    if excel == "":
        logging.info("没有输入<引物位置>.")
        bed = ""
    elif not excel.endswith(".xlsx"):
        logging.warning("输入<引物位置>文件不是.xlsx格式,请检查!")
        bed = ""
    else:
        gh = open(bed, "wt", encoding="utf-8", newline="")
        comments = ["参考序列名（不是文件名）", "chrom"]
        wb = load_workbook(excel)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=3):
            if row[0].value in comments:
                continue
            else:
                outline = "\t".join(list(map(lambda x:str(x.value), row)))
                gh.write(outline + "\n")
        gh.close()
    return bed

def zip_dir(src, dst):
    """
    zip压缩结果文件夹
    参数
        src     要压缩的文件夹
        dst     压缩后的文件
    """
    zip = zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED)
    for root, dirs, files in os.walk(src):
        arcpath = root.replace(os.path.dirname(src), "")
        for file in files:
            zip.write(os.path.join(root, file), arcname=os.path.join(arcpath, file))
    zip.close()

#VENUS专用######################################################################
def wrap_popen(shell, task_id, cursor, fh, timeout:int=604800):
    """[220602 mxf] 针对 VENUS 对 subprocess.Popen 的包装"""
    try:
        logging.debug(shell)
        pp = subprocess.Popen(shell, shell=True, encoding="utf-8", stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        popen_perftime = time.perf_counter()
        while True:
            time.sleep(0.1)
            pp.poll()
            if pp.returncode == 0:
                break
            elif (pp.returncode is not None) and (pp.returncode != 0):
                break
            # 获取taskid对应的状态, CANCEL则终止进程
            tmp_sql = f"select status from tb_analyse_task where task_id={task_id}"
            cursor.execute(tmp_sql)
            current_status = cursor.fetchone()["status"]
            if current_status == "CANCEL":
                pp.kill()
                break
            if time.perf_counter() - popen_perftime > timeout:  # 超过时间也会断掉
                pp.kill()
                break
        if pp.returncode is None:   # 运行记录 & Popen 返回码
            fh.write("终止运行!\n")
        else:
            fh.write("[STDOUT]\n" + pp.stdout.read())
            fh.write("\n[STDERR]\n" + pp.stderr.read())
        return pp.returncode
    except Exception as eh:
        print(eh)

def return2mysql(task_id, samples, cursor, return_fasta=True): 
    """
    [220601 mxf] 返回,每个样本都包括压缩包,html和fasta文件,.结果更新到MySQL数据库上, 对应任务ID和样本ID的返回位置
    一个样本一个报告
    """
    for name in samples: # 
        # 返回压缩包 
        return_file_path = f"/sdbb/Earth/AnalysisResults/{task_id}/{name}.zip"
        sql_path = f"update tb_task_sample set return_file_path='{return_file_path}' where task_id='{task_id}' and sample_number='{name}'"
        cursor.execute(sql_path)
        # 返回网页报告
        return_html_url = f"/file/weiyuan/AnalysisResults/{task_id}/{name}/index.html"
        sql_url = f"update tb_task_sample set return_html_url='{return_html_url}' where task_id='{task_id}' and sample_number='{name}'"
        cursor.execute(sql_url)
        # 返回fasta结果序列
        if return_fasta: # 可选传不传fasta
            return_fasta = f"/sdbb/Earth/AnalysisResults/{task_id}/{name}/{name}.fa"
            sql_fasta = f"update tb_task_sample set return_fasta='{return_fasta}' where task_id='{task_id}' and sample_number='{name}'"
            cursor.execute(sql_fasta)

def return2mysql_batch(task_id, samples, cursor): 
    """
    返回,每个样本都包括压缩包,html和fasta文件,.结果更新到MySQL数据库上, 对应任务ID和样本ID的返回位置
    一批样本一个报告
    """
    for name in samples: # 
        # 返回压缩包 
        return_file_path = f"/sdbb/Earth/AnalysisResults/{task_id}/{task_id}.zip"
        sql_path = f"update tb_task_sample set return_file_path='{return_file_path}' where task_id='{task_id}' and sample_number='{name}'"
        cursor.execute(sql_path)
        # 返回网页报告
        return_html_url = f"/file/weiyuan/AnalysisResults/{task_id}/index.html"
        sql_url = f"update tb_task_sample set return_html_url='{return_html_url}' where task_id='{task_id}' and sample_number='{name}'"
        cursor.execute(sql_url)

def lunx_upload(dir_analysis, dir_results, task_id, batch=False):
    """
    一个样本一个报告, Analysis 到 AnalysisResults 上载
        dir_analysis:   Analysis 样本目录
        dir_results:    AnalysisResults 样本目录
        batch:          布尔值,是否为一个批次一个报告, 比如溯源进化树
    """
    if os.path.isdir(dir_results): # 存在的话先删了，不然会 AnalysisResults/task_id/Upload
        shutil.rmtree(dir_results)
    run(f"cp -r {dir_analysis}/Upload {dir_results}", shell=True)
    # log文件压缩
    cml = f"cd {dir_analysis}\nzip -r {task_id}_log.zip logs"
    run(cml, shell=True)
    # 批次一个报告
    if batch:
        run(f"cp -r {dir_analysis}/{task_id}.zip {dir_results}", shell=True)
